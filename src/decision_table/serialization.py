"""Serialization: JSON, CSV, and Excel I/O for decision tables."""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from .model import (
    DONT_CARE,
    Action,
    Condition,
    ConditionType,
    Constraint,
    ConstraintType,
    DecisionTable,
    Range,
    Rule,
    TableType,
    parse_range,
)


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def save_json(table: DecisionTable, path: str | Path) -> None:
    """Save a decision table to a JSON file."""
    path = Path(path)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(table.to_dict(), f, indent=2, ensure_ascii=False)


def load_json(path: str | Path) -> DecisionTable:
    """Load a decision table from a JSON file."""
    path = Path(path)
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return DecisionTable.from_dict(data)


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

_ELSE_MARKER = "ELSE"
_META_PREFIX = "#"  # comment / metadata row prefix


def _rule_header(index: int, rule: Rule) -> str:
    """Build a column header for a rule."""
    label = f"Rule {index + 1}"
    if rule.is_else:
        label = f"Rule {index + 1} (ELSE)"
    return label


def _condition_type_tag(cond: Condition) -> str:
    """Return the Type column value for a condition row.

    For backward compatibility plain boolean conditions are just ``C``.
    Numeric and enum conditions carry the sub-type so they can be
    round-tripped.
    """
    if cond.condition_type == ConditionType.BOOLEAN:
        return "C"
    elif cond.condition_type == ConditionType.NUMERIC:
        return "C:NUMERIC"
    else:
        return "C:ENUM"


def _parse_type_tag(tag: str) -> tuple[str, ConditionType | None]:
    """Parse a Type column value back into (base_type, condition_type | None).

    Returns ``("C", <ConditionType>)`` for condition rows and
    ``("A", None)`` for action rows, etc.
    """
    tag = tag.strip().upper()
    if tag.startswith("C:"):
        sub = tag.split(":", 1)[1]
        ct_map = {"NUMERIC": ConditionType.NUMERIC, "ENUM": ConditionType.ENUM}
        return "C", ct_map.get(sub, None)
    if tag == "C":
        return "C", ConditionType.BOOLEAN
    return tag, None


# ---------------------------------------------------------------------------
# CSV save / load
# ---------------------------------------------------------------------------

def save_csv(table: DecisionTable, path: str | Path) -> None:
    """Save a decision table to a CSV file.

    Format::

        # table_type=single_hit
        # constraint: exclusion | CondA=x, CondB=y | description text
        Type, Name, Rule 1, Rule 2, Rule 3 (ELSE)
        # priority, , 10, 5, -1
        C, CondName, T, F, -
        C:NUMERIC, Age, <18, 18-64, -
        A, ActName, X, , fire
    """
    path = Path(path)
    num_rules = len(table.rules)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # -- Metadata comment rows --
        writer.writerow([_META_PREFIX, f"table_type={table.table_type.value}"])

        # Constraints
        for con in table.constraints:
            cond_str = ", ".join(f"{k}={v}" for k, v in con.conditions.items())
            writer.writerow([
                _META_PREFIX,
                f"constraint: {con.constraint_type.value} | {cond_str} | {con.description}",
            ])

        # -- Header row --
        rule_headers = [_rule_header(i, r) for i, r in enumerate(table.rules)]
        writer.writerow(["Type", "Name"] + rule_headers)

        # -- Priority row (only if any rule has a non-zero priority) --
        if any(r.priority != 0 for r in table.rules):
            writer.writerow(
                [_META_PREFIX, "priority"]
                + [str(r.priority) for r in table.rules]
            )

        # -- Condition rows --
        for cond in table.conditions:
            tag = _condition_type_tag(cond)
            values: list[str] = []
            for rule in table.rules:
                if rule.is_else:
                    values.append(_ELSE_MARKER)
                else:
                    values.append(rule.condition_entries.get(cond.name, DONT_CARE))
            writer.writerow([tag, cond.name] + values)

        # -- Action rows --
        for action in table.actions:
            values = [rule.action_entries.get(action.name, "") for rule in table.rules]
            writer.writerow(["A", action.name] + values)


def load_csv(path: str | Path) -> DecisionTable:
    """Load a decision table from a CSV file.

    Backward compatible: files without metadata rows or condition-type
    annotations are loaded using the original heuristic.
    """
    path = Path(path)
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return DecisionTable()

    # ---- Separate metadata comment rows from data rows ----
    meta_rows: list[list[str]] = []
    data_rows: list[list[str]] = []

    for row in rows:
        if not row:
            continue
        first = row[0].strip()
        if first == _META_PREFIX:
            meta_rows.append(row)
        else:
            data_rows.append(row)

    # ---- Parse metadata ----
    table_type = TableType.SINGLE_HIT
    constraints: list[Constraint] = []
    priority_values: list[int] | None = None

    for mrow in meta_rows:
        payload = mrow[1].strip() if len(mrow) > 1 else ""

        # table_type
        if payload.startswith("table_type="):
            try:
                table_type = TableType(payload.split("=", 1)[1].strip())
            except ValueError:
                pass

        # constraint
        elif payload.startswith("constraint:"):
            _parse_constraint_meta(payload, constraints)

        # priority
        elif payload == "priority":
            priority_values = [int(v.strip()) if v.strip() else 0 for v in mrow[2:]]

    # ---- Skip the header row if present ----
    has_header = False
    if data_rows and data_rows[0][0].strip().lower() in ("type",):
        header_row = data_rows.pop(0)
        has_header = True

    # ---- Detect else columns from header ----
    else_columns: set[int] = set()
    if has_header:
        for ci, hval in enumerate(header_row[2:]):
            if _ELSE_MARKER in hval.upper():
                else_columns.add(ci)

    # ---- Parse condition / action rows ----
    conditions: list[Condition] = []
    actions: list[Action] = []
    condition_values: dict[str, list[str]] = {}
    action_values: dict[str, list[str]] = {}
    condition_types: dict[str, ConditionType] = {}
    num_rules = 0

    for row in data_rows:
        if len(row) < 2:
            continue
        base_type, cond_type = _parse_type_tag(row[0])
        name = row[1].strip()
        values = [v.strip() for v in row[2:]]
        num_rules = max(num_rules, len(values))

        if base_type == "C":
            # For else columns the cell value is ELSE; treat as don't-care for
            # determining possible_values.
            clean_values = [
                v for i, v in enumerate(values)
                if v and v != DONT_CARE and v.upper() != _ELSE_MARKER
            ]
            unique_vals = sorted(set(clean_values))

            if cond_type == ConditionType.NUMERIC:
                # Attempt to parse range labels
                ranges: list[Range] = []
                for lbl in unique_vals:
                    try:
                        ranges.append(parse_range(lbl))
                    except (ValueError, IndexError):
                        pass
                conditions.append(Condition(
                    name=name,
                    possible_values=unique_vals if unique_vals else ["T", "F"],
                    condition_type=ConditionType.NUMERIC,
                    ranges=ranges,
                ))
            elif cond_type == ConditionType.ENUM:
                conditions.append(Condition(
                    name=name,
                    possible_values=unique_vals if unique_vals else ["T", "F"],
                    condition_type=ConditionType.ENUM,
                ))
            else:
                # Default / boolean
                if not unique_vals:
                    unique_vals = ["T", "F"]
                conditions.append(Condition(name=name, possible_values=unique_vals))

            condition_types[name] = cond_type or ConditionType.BOOLEAN
            condition_values[name] = values

        elif base_type == "A":
            unique_vals = sorted({v for v in values if v})
            if not unique_vals:
                unique_vals = ["X", ""]
            if "" not in unique_vals:
                unique_vals.append("")
            actions.append(Action(name=name, possible_values=unique_vals))
            action_values[name] = values

    # ---- Also detect else columns from cell values ----
    for i in range(num_rules):
        for cond in conditions:
            vals = condition_values.get(cond.name, [])
            if i < len(vals) and vals[i].upper() == _ELSE_MARKER:
                else_columns.add(i)

    # ---- Build rules ----
    rules: list[Rule] = []
    for i in range(num_rules):
        is_else = i in else_columns

        cond_entries: dict[str, str] = {}
        for cond in conditions:
            vals = condition_values[cond.name]
            raw = vals[i] if i < len(vals) else DONT_CARE
            if raw.upper() == _ELSE_MARKER:
                raw = DONT_CARE
            cond_entries[cond.name] = raw

        act_entries: dict[str, str] = {}
        for action in actions:
            vals = action_values[action.name]
            act_entries[action.name] = vals[i] if i < len(vals) else ""

        priority = 0
        if priority_values and i < len(priority_values):
            priority = priority_values[i]

        rules.append(Rule(
            condition_entries=cond_entries,
            action_entries=act_entries,
            priority=priority,
            is_else=is_else,
        ))

    table = DecisionTable(
        name=Path(path).stem,
        conditions=conditions,
        actions=actions,
        rules=rules,
        constraints=constraints,
        table_type=table_type,
    )
    return table


def _parse_constraint_meta(payload: str, constraints: list[Constraint]) -> None:
    """Parse a ``constraint: type | conds | desc`` metadata string."""
    # "constraint: exclusion | CondA=x, CondB=y | some description"
    rest = payload.split(":", 1)[1].strip()
    parts = [p.strip() for p in rest.split("|")]
    if len(parts) < 2:
        return
    try:
        ct = ConstraintType(parts[0].strip())
    except ValueError:
        return
    cond_pairs: dict[str, str] = {}
    for pair in parts[1].split(","):
        pair = pair.strip()
        if "=" in pair:
            k, v = pair.split("=", 1)
            cond_pairs[k.strip()] = v.strip()
    desc = parts[2].strip() if len(parts) > 2 else ""
    constraints.append(Constraint(constraint_type=ct, conditions=cond_pairs, description=desc))


# ---------------------------------------------------------------------------
# Excel save / load
# ---------------------------------------------------------------------------

def save_excel(table: DecisionTable, path: str | Path) -> None:
    """Save a decision table to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = (table.name or "Decision Table")[:31]

    header_font = Font(bold=True)
    cond_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    action_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    else_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    meta_font = Font(italic=True, color="808080")
    center = Alignment(horizontal="center")

    current_row = 1

    # -- Metadata row: table type --
    ws.cell(row=current_row, column=1, value=_META_PREFIX).font = meta_font
    ws.cell(row=current_row, column=2, value=f"table_type={table.table_type.value}").font = meta_font
    current_row += 1

    # -- Constraint rows --
    for con in table.constraints:
        cond_str = ", ".join(f"{k}={v}" for k, v in con.conditions.items())
        ws.cell(row=current_row, column=1, value=_META_PREFIX).font = meta_font
        ws.cell(
            row=current_row, column=2,
            value=f"constraint: {con.constraint_type.value} | {cond_str} | {con.description}",
        ).font = meta_font
        current_row += 1

    # -- Header row --
    ws.cell(row=current_row, column=1, value="Type").font = header_font
    ws.cell(row=current_row, column=2, value="Name").font = header_font
    for i, rule in enumerate(table.rules):
        cell = ws.cell(row=current_row, column=3 + i, value=_rule_header(i, rule))
        cell.font = header_font
        cell.alignment = center
    current_row += 1

    # -- Priority row --
    if any(r.priority != 0 for r in table.rules):
        ws.cell(row=current_row, column=1, value=_META_PREFIX).font = meta_font
        ws.cell(row=current_row, column=2, value="priority").font = meta_font
        for i, rule in enumerate(table.rules):
            ws.cell(row=current_row, column=3 + i, value=str(rule.priority)).font = meta_font
        current_row += 1

    # -- Condition rows --
    for cond in table.conditions:
        tag = _condition_type_tag(cond)
        ws.cell(row=current_row, column=1, value=tag).fill = cond_fill
        ws.cell(row=current_row, column=2, value=cond.name).fill = cond_fill
        for i, rule in enumerate(table.rules):
            if rule.is_else:
                val = _ELSE_MARKER
                fill = else_fill
            else:
                val = rule.condition_entries.get(cond.name, DONT_CARE)
                fill = cond_fill
            cell = ws.cell(row=current_row, column=3 + i, value=val)
            cell.alignment = center
            cell.fill = fill
        current_row += 1

    # -- Action rows --
    for action in table.actions:
        ws.cell(row=current_row, column=1, value="A").fill = action_fill
        ws.cell(row=current_row, column=2, value=action.name).fill = action_fill
        for i, rule in enumerate(table.rules):
            val = rule.action_entries.get(action.name, "")
            fill = else_fill if rule.is_else else action_fill
            cell = ws.cell(row=current_row, column=3 + i, value=val)
            cell.alignment = center
            cell.fill = fill
        current_row += 1

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 8)

    wb.save(path)


def load_excel(path: str | Path) -> DecisionTable:
    """Load a decision table from an Excel file."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v is not None else "" for v in row])
    wb.close()

    if not rows:
        return DecisionTable()

    # ---- Separate metadata comment rows from data rows ----
    meta_rows: list[list[str]] = []
    data_rows: list[list[str]] = []

    for row in rows:
        if not row:
            continue
        first = row[0].strip()
        if first == _META_PREFIX:
            meta_rows.append(row)
        else:
            data_rows.append(row)

    # ---- Parse metadata (same logic as CSV) ----
    table_type = TableType.SINGLE_HIT
    constraints: list[Constraint] = []
    priority_values: list[int] | None = None

    for mrow in meta_rows:
        payload = mrow[1].strip() if len(mrow) > 1 else ""

        if payload.startswith("table_type="):
            try:
                table_type = TableType(payload.split("=", 1)[1].strip())
            except ValueError:
                pass
        elif payload.startswith("constraint:"):
            _parse_constraint_meta(payload, constraints)
        elif payload == "priority":
            priority_values = [int(v.strip()) if v.strip() else 0 for v in mrow[2:]]

    # ---- Skip header row if present ----
    has_header = False
    header_row: list[str] = []
    if data_rows and data_rows[0][0].strip().lower() in ("type",):
        header_row = data_rows.pop(0)
        has_header = True

    else_columns: set[int] = set()
    if has_header:
        for ci, hval in enumerate(header_row[2:]):
            if _ELSE_MARKER in hval.upper():
                else_columns.add(ci)

    # ---- Parse condition / action rows ----
    conditions: list[Condition] = []
    actions: list[Action] = []
    condition_values: dict[str, list[str]] = {}
    action_values: dict[str, list[str]] = {}
    condition_types: dict[str, ConditionType] = {}
    num_rules = 0

    for row in data_rows:
        if len(row) < 2:
            continue
        base_type, cond_type = _parse_type_tag(row[0])
        name = row[1].strip()
        values = [v.strip() for v in row[2:]]
        num_rules = max(num_rules, len(values))

        if base_type == "C":
            clean_values = [
                v for i, v in enumerate(values)
                if v and v != DONT_CARE and v.upper() != _ELSE_MARKER
            ]
            unique_vals = sorted(set(clean_values))

            if cond_type == ConditionType.NUMERIC:
                ranges: list[Range] = []
                for lbl in unique_vals:
                    try:
                        ranges.append(parse_range(lbl))
                    except (ValueError, IndexError):
                        pass
                conditions.append(Condition(
                    name=name,
                    possible_values=unique_vals if unique_vals else ["T", "F"],
                    condition_type=ConditionType.NUMERIC,
                    ranges=ranges,
                ))
            elif cond_type == ConditionType.ENUM:
                conditions.append(Condition(
                    name=name,
                    possible_values=unique_vals if unique_vals else ["T", "F"],
                    condition_type=ConditionType.ENUM,
                ))
            else:
                if not unique_vals:
                    unique_vals = ["T", "F"]
                conditions.append(Condition(name=name, possible_values=unique_vals))

            condition_types[name] = cond_type or ConditionType.BOOLEAN
            condition_values[name] = values

        elif base_type == "A":
            unique_vals = sorted({v for v in values if v})
            if not unique_vals:
                unique_vals = ["X", ""]
            if "" not in unique_vals:
                unique_vals.append("")
            actions.append(Action(name=name, possible_values=unique_vals))
            action_values[name] = values

    # Detect else columns from cell values
    for i in range(num_rules):
        for cond in conditions:
            vals = condition_values.get(cond.name, [])
            if i < len(vals) and vals[i].upper() == _ELSE_MARKER:
                else_columns.add(i)

    # ---- Build rules ----
    rules: list[Rule] = []
    for i in range(num_rules):
        is_else = i in else_columns

        cond_entries: dict[str, str] = {}
        for cond in conditions:
            vals = condition_values[cond.name]
            raw = vals[i] if i < len(vals) else DONT_CARE
            if raw.upper() == _ELSE_MARKER:
                raw = DONT_CARE
            cond_entries[cond.name] = raw

        act_entries: dict[str, str] = {}
        for action in actions:
            vals = action_values[action.name]
            act_entries[action.name] = vals[i] if i < len(vals) else ""

        priority = 0
        if priority_values and i < len(priority_values):
            priority = priority_values[i]

        rules.append(Rule(
            condition_entries=cond_entries,
            action_entries=act_entries,
            priority=priority,
            is_else=is_else,
        ))

    table = DecisionTable(
        name=Path(path).stem,
        conditions=conditions,
        actions=actions,
        rules=rules,
        constraints=constraints,
        table_type=table_type,
    )
    return table


# ---------------------------------------------------------------------------
# Test-case export helpers
# ---------------------------------------------------------------------------

def save_test_cases_csv(
    test_cases: list[Any],
    table: DecisionTable,
    path: str | Path,
) -> None:
    """Export test cases to a CSV file.

    Each row contains the test number, type, description, all condition
    input values, all expected action output values, and covering rules.
    """
    path = Path(path)
    cond_names = [c.name for c in table.conditions]
    act_names = [a.name for a in table.actions]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        header = ["#", "Type", "Description"] + cond_names + act_names + ["Covering Rules"]
        writer.writerow(header)
        for i, tc in enumerate(test_cases, 1):
            inputs = tc.inputs if hasattr(tc, "inputs") else tc.get("inputs", {})
            outputs = (
                tc.expected_outputs
                if hasattr(tc, "expected_outputs")
                else tc.get("expected_outputs", {})
            )
            test_type = tc.test_type if hasattr(tc, "test_type") else tc.get("test_type", "")
            description = tc.description if hasattr(tc, "description") else tc.get("description", "")
            covering = (
                tc.covering_rules
                if hasattr(tc, "covering_rules")
                else tc.get("covering_rules", [])
            )

            row: list[Any] = [i, test_type, description]
            row.extend(inputs.get(c, "") for c in cond_names)
            row.extend(outputs.get(a, "") for a in act_names)
            row.append(", ".join(f"R{r + 1}" for r in covering))
            writer.writerow(row)


def save_test_cases_excel(
    test_cases: list[Any],
    table: DecisionTable,
    path: str | Path,
) -> None:
    """Export test cases to an Excel file with formatting.

    Condition-input columns get a blue tint, expected-action columns a
    green tint, and boundary / else test rows are highlighted.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    header_font = Font(bold=True)
    cond_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    action_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    boundary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    else_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    center = Alignment(horizontal="center")

    cond_names = [c.name for c in table.conditions]
    act_names = [a.name for a in table.actions]

    # Header
    headers = ["#", "Type", "Description"] + cond_names + act_names + ["Covering Rules"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.alignment = center

    # Data rows
    for ri, tc in enumerate(test_cases, 1):
        inputs = tc.inputs if hasattr(tc, "inputs") else tc.get("inputs", {})
        outputs = (
            tc.expected_outputs
            if hasattr(tc, "expected_outputs")
            else tc.get("expected_outputs", {})
        )
        test_type = tc.test_type if hasattr(tc, "test_type") else tc.get("test_type", "")
        description = tc.description if hasattr(tc, "description") else tc.get("description", "")
        covering = (
            tc.covering_rules
            if hasattr(tc, "covering_rules")
            else tc.get("covering_rules", [])
        )

        excel_row = ri + 1
        ws.cell(row=excel_row, column=1, value=ri).alignment = center
        ws.cell(row=excel_row, column=2, value=test_type).alignment = center
        ws.cell(row=excel_row, column=3, value=description)

        # Condition input cells
        col_offset = 4
        for ci, cname in enumerate(cond_names):
            cell = ws.cell(row=excel_row, column=col_offset + ci, value=inputs.get(cname, ""))
            cell.alignment = center
            cell.fill = cond_fill

        # Action output cells
        col_offset = 4 + len(cond_names)
        for ai, aname in enumerate(act_names):
            cell = ws.cell(row=excel_row, column=col_offset + ai, value=outputs.get(aname, ""))
            cell.alignment = center
            cell.fill = action_fill

        # Covering rules
        ws.cell(
            row=excel_row,
            column=4 + len(cond_names) + len(act_names),
            value=", ".join(f"R{r + 1}" for r in covering),
        ).alignment = center

        # Row-level highlight for special test types
        if test_type == "boundary":
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).fill = boundary_fill
        elif test_type == "else":
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).fill = else_fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 8)

    wb.save(path)


# ---------------------------------------------------------------------------
# Generic dispatch by extension
# ---------------------------------------------------------------------------

def save_file(table: DecisionTable, path: str | Path) -> None:
    """Save a decision table, choosing format by file extension."""
    path = Path(path)
    ext = path.suffix.lower()
    if ext == ".json":
        save_json(table, path)
    elif ext == ".csv":
        save_csv(table, path)
    elif ext in (".xlsx", ".xls"):
        save_excel(table, path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def load_file(path: str | Path) -> DecisionTable:
    """Load a decision table, choosing format by file extension."""
    path = Path(path)
    ext = path.suffix.lower()
    if ext == ".json":
        return load_json(path)
    elif ext == ".csv":
        return load_csv(path)
    elif ext in (".xlsx", ".xls"):
        return load_excel(path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")
